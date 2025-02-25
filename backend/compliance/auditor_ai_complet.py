# -*- coding: utf-8 -*-
"""
Created on Thu Sep 26 21:43:07 2024

@author: Daniel Lopez
"""

from ollama import generate
from opensearchpy import OpenSearch, exceptions as opensearch_exceptions
import os
import docx
import openpyxl
from PyPDF2 import PdfReader
import time
import tqdm  # Para la barra de progreso
import warnings
from urllib3.exceptions import InsecureRequestWarning

# Ignorar advertencias de solicitudes inseguras (verificación SSL desactivada)
warnings.simplefilter('ignore', InsecureRequestWarning)

# Conexión a OpenSearch con verificación de SSL desactivada
def connect_to_opensearch():
    host = 'https://localhost:9201'  # Dirección del servidor de OpenSearch
    auth = ('admin', 'Soporte18*')  # Credenciales de OpenSearch (usuario, contraseña)
    
    try:
        client = OpenSearch(
            hosts=[host],
            http_auth=auth,
            use_ssl=True,
            verify_certs=False,
            timeout=120
        )
        return client
    except Exception as e:
        print(f"Error al conectar con OpenSearch: {e}")
        return None

# Obtención del texto de la sección desde OpenSearch
def obtener_texto_seccion(seccion):
    client = connect_to_opensearch()
    if client is None:
        return None
    
    try:
        response = client.search(
            index="soa_iso_27001",
            body={
                "query": {
                    "match": {
                        "Sección": seccion
                    }
                }
            }
        )
        hits = response['hits']['hits']
        if hits:
            resultado = hits[0]['_source']
            return {
                'descripcion': resultado.get('Descripción del Control', 'No se encontró la descripción'),
                'objetivo_control': resultado.get('Objetivo de Control / Control', 'No se encontró el objetivo del control'),
                'evidencia': resultado.get('Soporte/Evidencia', 'No se encontró evidencia')
            }
        else:
            print(f"ALERTA: La sección {seccion} no se encuentra en el índice.")
            return None
    except Exception as e:
        print(f"Error al buscar la sección en OpenSearch: {e}")
        return None

# Barra de progreso simulada
def mostrar_barra_progreso(duracion):
    for _ in tqdm.tqdm(range(duracion), desc="Validando el documento"):
        time.sleep(0.1)

# Función para validar con Ollama
def validar_con_ollama(contenido_documento, seccion, objetivo_control, descripcion_control, evidencia):
    prompt = f""" que entiendes del control {seccion} y el objetivo de control  {objetivo_control} descripcion_control {descripcion_control} Y el con el contenido del documento {contenido_documento} complemente la respuesta  
    Estructura de la respuesta
    1. Entendimiento del objetivo de control
    2. Recomendaciones de implementación y evidencias sugeridas
    """
    
    
    try:
        mostrar_barra_progreso(50)
        resultado = generate(model="llama3.1", prompt=prompt)
        return resultado  # Retornamos el resultado completo
    except Exception as e:
        print(f"Error al generar la validación con Ollama: {e}")
        return None

# Indexación del resultado en OpenSearch
def indexar_resultado_opensearch(client, seccion, resultado):
    if not resultado:
        print(f"ALERTA: No se puede indexar un resultado vacío para la sección {seccion}.")
        return
    
    doc = {'seccion': seccion, 'resultado': resultado}
    
    try:
        response = client.index(index='iso27001_validaciones', body=doc)
        print(f"Resultado indexado en OpenSearch: {response}")
    except opensearch_exceptions.ConnectionError as ce:
        print(f"Error de conexión con OpenSearch: {ce}")
    except opensearch_exceptions.RequestError as re:
        print(f"Error en la solicitud a OpenSearch: {re}")
    except Exception as e:
        print(f"Error inesperado al indexar en OpenSearch: {e}")

# Lectura del contenido de un documento según su tipo
def leer_contenido_documento(ruta_documento):
    extension = os.path.splitext(ruta_documento)[1].lower()
    try:
        if extension == '.docx':
            return leer_docx(ruta_documento)
        elif extension == '.xlsx':
            return leer_xlsx(ruta_documento)
        elif extension == '.pdf':
            return leer_pdf(ruta_documento)
        else:
            print(f"Tipo de archivo {extension} no soportado.")
            return None
    except Exception as e:
        print(f"Error al leer el documento {ruta_documento}: {e}")
        return None

# Lectura de archivos .docx
def leer_docx(ruta_documento):
    doc = docx.Document(ruta_documento)
    return '\n'.join([paragraph.text for paragraph in doc.paragraphs])

# Lectura de archivos .xlsx
def leer_xlsx(ruta_documento):
    workbook = openpyxl.load_workbook(ruta_documento)
    contenido = []
    for sheet in workbook.sheetnames:
        worksheet = workbook[sheet]
        for row in worksheet.iter_rows(values_only=True):
            contenido.append(' '.join([str(cell) for cell in row if cell]))
    return '\n'.join(contenido)

# Lectura de archivos .pdf
def leer_pdf(ruta_documento):
    reader = PdfReader(ruta_documento)
    return '\n'.join([page.extract_text() for page in reader.pages])

# Validación del contenido del documento según la sección
def validar_contenido_seccion(ruta_documento, seccion):
    contenido_documento = leer_contenido_documento(ruta_documento)
    
    if contenido_documento:
        datos_seccion = obtener_texto_seccion(seccion)
        if datos_seccion:
            print(f"Sección seleccionada: {seccion}")
            print(f"Descripción del Control: {datos_seccion['descripcion']}")
            print(f"Objetivo del Control: {datos_seccion['objetivo_control']}")
            print(f"Evidencia Esperada: {datos_seccion['evidencia']}")
            
            # Pasar tanto el objetivo de control, la descripción del control y la evidencia
            resultado_ollama = validar_con_ollama(
                contenido_documento,
                seccion,
                datos_seccion['objetivo_control'],
                datos_seccion['descripcion'],
                datos_seccion['evidencia']
            )
            
            if resultado_ollama:
                print(f"Resultado de la validación para la sección {seccion}:\n{resultado_ollama}")
                
                client = connect_to_opensearch()
                if client:
                    indexar_resultado_opensearch(client, seccion, resultado_ollama)
            else:
                print(f"ALERTA: La validación con Ollama falló para la sección {seccion}.")
        else:
            print(f"ALERTA: La sección {seccion} no se encuentra.")
    else:
        print(f"ALERTA: No se proporcionó contenido para la validación de la sección {seccion}.")

# Solicitar entrada de texto para la sección y el documento
seccion = input("Ingrese la sección ISO 27001 que desea validar (ejemplo: 5.1): ")

# Ruta donde están los documentos
directorio = 'C:/Users/Daniel Lopez/Documents/Documentos_sgsi'

# Listar todos los archivos en el directorio
archivos = os.listdir(directorio)
print("Documentos disponibles:")
for idx, archivo in enumerate(archivos, 1):
    print(f"{idx}. {archivo}")

# Solicitar el número del documento a validar
eleccion = int(input("Seleccione el número del documento que desea validar: "))
ruta_documento = os.path.join(directorio, archivos[eleccion - 1])

# Llamar a la función para validar el contenido del documento desde la ruta proporcionada
validar_contenido_seccion(ruta_documento, seccion)
